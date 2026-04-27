"""
excel_writer.py
Genera el archivo Excel consolidado a partir de la lista de facturas procesadas.
Aplica formato profesional y garantiza tipos de datos correctos.
"""
 
import logging
from io import BytesIO
from typing import List
 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
 
from app.parser.extractor import COLUMNAS_SALIDA
from app.parser.normalizer import CAMPOS_NUMERICOS, CAMPOS_FECHA, CAMPOS_ENTEROS
 
logger = logging.getLogger(__name__)
 
 
# ──────────────────────────────────────────────────────────────
# Configuración visual
# ──────────────────────────────────────────────────────────────
 
COLOR_ENCABEZADO_BG = "1F3864"   # Azul oscuro
COLOR_ENCABEZADO_FX = "FFFFFF"   # Blanco
COLOR_FILA_PAR      = "EEF2F7"   # Gris muy claro
COLOR_ERROR_BG      = "FFE0E0"   # Rosa claro para errores
COLOR_BORDE         = "C0C0C0"   # Gris claro
 
ETIQUETAS_COLUMNAS = {
    "archivo":                              "archivo",
    "nro_cuenta":                           "nro_cuenta",
    "nro_factura":                          "nro_factura",
    "fecha_emision":                        "fecha_emision",
    "prox_vencimiento":                     "prox_vencimiento",
    "acuerdo_servicio":                     "acuerdo_servicio",
    "nro_medidor":                          "nro_medidor",
    "tarifa_aplicada":                      "tarifa_aplicada",
    "nombre_cliente":                       "nombre_cliente",
    "direccion_cliente":                    "direccion_cliente",
    "localidad_cliente":                    "localidad_cliente",
    "departamento_cliente":                 "departamento_cliente",
    "potencia_contratada_punta_llano_kw":   "potencia_contratada_punta_llano_kw",
    "potencia_contratada_valle_kw":         "potencia_contratada_valle_kw",
    "consumo_activo_kwh":                   "consumo_activo_kwh",
    "consumo_reactivo_kvarh":               "consumo_reactivo_kvarh",
    "tension":                              "tension",
    "fases":                                "fases",
    "direccion_servicio":                   "direccion_servicio",
    "periodo_consumo":                      "periodo_consumo",
    "zona_electrica":                       "zona_electrica",
    "total_detalle_facturacion":            "total_detalle_facturacion",
    "energa_llano_factor":                  "energa_llano_factor",
    "energa_llano_lect_act":                "energa_llano_lect_act",
    "energa_llano_lect_ant":                "energa_llano_lect_ant",
    "energa_llano_tipo_lec":                "energa_llano_tipo_lec",
    "energa_llano_total":                   "energa_llano_total",
    "energa_punta_factor":                  "energa_punta_factor",
    "energa_punta_lect_act":                "energa_punta_lect_act",
    "energa_punta_lect_ant":                "energa_punta_lect_ant",
    "energa_punta_tipo_lec":                "energa_punta_tipo_lec",
    "energa_punta_total":                   "energa_punta_total",
    "energa_reactiva_factor":               "energa_reactiva_factor",
    "energa_reactiva_lect_act":             "energa_reactiva_lect_act",
    "energa_reactiva_lect_ant":             "energa_reactiva_lect_ant",
    "energa_reactiva_tipo_lec":             "energa_reactiva_tipo_lec",
    "energa_reactiva_total":                "energa_reactiva_total",
    "energa_sal_llano_factor":              "energa_sal_llano_factor",
    "energa_sal_llano_lect_act":            "energa_sal_llano_lect_act",
    "energa_sal_llano_lect_ant":            "energa_sal_llano_lect_ant",
    "energa_sal_llano_tipo_lec":            "energa_sal_llano_tipo_lec",
    "energa_sal_llano_total":               "energa_sal_llano_total",
    "energa_sal_punta_factor":              "energa_sal_punta_factor",
    "energa_sal_punta_lect_act":            "energa_sal_punta_lect_act",
    "energa_sal_punta_lect_ant":            "energa_sal_punta_lect_ant",
    "energa_sal_punta_tipo_lec":            "energa_sal_punta_tipo_lec",
    "energa_sal_punta_total":               "energa_sal_punta_total",
    "energa_sal_valle_factor":              "energa_sal_valle_factor",
    "energa_sal_valle_lect_act":            "energa_sal_valle_lect_act",
    "energa_sal_valle_lect_ant":            "energa_sal_valle_lect_ant",
    "energa_sal_valle_tipo_lec":            "energa_sal_valle_tipo_lec",
    "energa_sal_valle_total":               "energa_sal_valle_total",
    "energa_valle_factor":                  "energa_valle_factor",
    "energa_valle_lect_act":                "energa_valle_lect_act",
    "energa_valle_lect_ant":                "energa_valle_lect_ant",
    "energa_valle_tipo_lec":                "energa_valle_tipo_lec",
    "energa_valle_total":                   "energa_valle_total",
    "npags_pdf":                            "npags_pdf",
    "potencia_factor":                      "potencia_factor",
    "potencia_lect_act":                    "potencia_lect_act",
    "potencia_lect_ant":                    "potencia_lect_ant",
    "potencia_tipo_lec":                    "potencia_tipo_lec",
    "potencia_total":                       "potencia_total",
    "potencia_valle_factor":                "potencia_valle_factor",
    "potencia_valle_lect_act":              "potencia_valle_lect_act",
    "potencia_valle_lect_ant":              "potencia_valle_lect_ant",
    "potencia_valle_tipo_lec":              "potencia_valle_tipo_lec",
    "potencia_valle_total":                 "potencia_valle_total",
}
 
 
# ──────────────────────────────────────────────────────────────
# Función principal
# ──────────────────────────────────────────────────────────────
 
def generar_excel(registros: List[dict]) -> bytes:
    """
    Genera el archivo Excel consolidado en memoria (BytesIO).
 
    Args:
        registros: Lista de diccionarios, uno por factura procesada.
 
    Returns:
        Bytes del archivo .xlsx listo para descarga.
    """
    if not registros:
        logger.warning("No hay registros para generar el Excel.")
 
    # ── 1. Construir DataFrame con columnas exactas ──────────────
    df = _construir_dataframe(registros)
 
    # ── 2. Guardar con pandas a BytesIO ─────────────────────────
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Facturas", index=False, header=False, startrow=1)
 
    buffer.seek(0)
 
    # ── 3. Reabrir con openpyxl para formatear ───────────────────
    wb = load_workbook(buffer)
    ws = wb.active
    ws.title = "Facturas"
 
    _escribir_encabezados(ws, df.columns.tolist())
    _formatear_filas(ws, df)
    _ajustar_anchos(ws, df)
    _agregar_hoja_resumen(wb, df)
 
    # ── 4. Serializar y retornar bytes ───────────────────────────
    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida.read()
 
 
# ──────────────────────────────────────────────────────────────
# Construcción del DataFrame
# ──────────────────────────────────────────────────────────────
 
def _construir_dataframe(registros: List[dict]) -> pd.DataFrame:
    """
    Construye el DataFrame con el orden exacto de columnas y tipos correctos.
    Los campos numéricos se fuerzan a tipo numérico; el resto como string.
    """
    # Filtrar columnas internas (prefijadas con _)
    columnas_validas = [c for c in COLUMNAS_SALIDA]
 
    filas = []
    for reg in registros:
        fila = {}
        for col in columnas_validas:
            fila[col] = reg.get(col)
        filas.append(fila)
 
    df = pd.DataFrame(filas, columns=columnas_validas)
 
    # Forzar tipos numéricos
    for col in CAMPOS_NUMERICOS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
 
    for col in CAMPOS_ENTEROS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
 
    # Usar etiquetas legibles como nombres de columna
    df.columns = [ETIQUETAS_COLUMNAS.get(c, c) for c in df.columns]
 
    return df
 
 
# ──────────────────────────────────────────────────────────────
# Formateo del Excel
# ──────────────────────────────────────────────────────────────
 
def _escribir_encabezados(ws, columnas: list):
    """Escribe y formatea la fila de encabezados."""
    font_enc = Font(name="Arial", bold=True, color=COLOR_ENCABEZADO_FX, size=10)
    fill_enc = PatternFill("solid", fgColor=COLOR_ENCABEZADO_BG)
    alin_enc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = _borde_fino()
 
    for col_idx, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.font = font_enc
        celda.fill = fill_enc
        celda.alignment = alin_enc
        celda.border = borde
 
    ws.row_dimensions[1].height = 36
 
 
def _formatear_filas(ws, df: pd.DataFrame):
    """Aplica formato alternado a las filas de datos."""
    font_datos = Font(name="Arial", size=9)
    fill_par   = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    fill_error = PatternFill("solid", fgColor=COLOR_ERROR_BG)
    alin_centro = Alignment(horizontal="center", vertical="center")
    alin_izq    = Alignment(horizontal="left", vertical="center")
    alin_der    = Alignment(horizontal="right", vertical="center")
    borde = _borde_fino()
 
    # Columnas numéricas por índice (1-based)
    etiquetas_num = {ETIQUETAS_COLUMNAS.get(c, c) for c in CAMPOS_NUMERICOS | CAMPOS_ENTEROS}
    idx_numericos = {
        i + 1 for i, col in enumerate(df.columns)
        if col in etiquetas_num
    }
 
    for fila_idx, (_, fila) in enumerate(df.iterrows(), start=2):
        es_par = (fila_idx % 2 == 0)
        fill_fila = fill_par if es_par else None
 
        for col_idx in range(1, len(df.columns) + 1):
            celda = ws.cell(row=fila_idx, column=col_idx)
            celda.font = font_datos
            celda.border = borde
 
            if fill_fila:
                celda.fill = fill_fila
 
            # Alineación según tipo de columna
            if col_idx in idx_numericos:
                celda.alignment = alin_der
                if isinstance(celda.value, (int, float)):
                    celda.number_format = '#,##0.00'
            else:
                celda.alignment = alin_izq
 
 
def _ajustar_anchos(ws, df: pd.DataFrame):
    """Ajusta el ancho de columnas según el contenido."""
    ANCHOS_MINIMOS = {"Archivo": 30, "Nombre Cliente": 28, "Dirección Cliente": 32}
    ANCHO_DEFAULT_MAX = 25
 
    for col_idx, col_nombre in enumerate(df.columns, start=1):
        letra = get_column_letter(col_idx)
        ancho_header = len(str(col_nombre)) + 2
 
        # Calcular ancho por datos
        ancho_datos = 0
        for valor in df[col_nombre]:
            if valor is not None and not pd.isna(valor):
                ancho_datos = max(ancho_datos, len(str(valor)))
 
        ancho_final = max(ancho_header, min(ancho_datos + 2, ANCHO_DEFAULT_MAX))
        ancho_final = max(ancho_final, ANCHOS_MINIMOS.get(col_nombre, 10))
 
        ws.column_dimensions[letra].width = ancho_final
 
    # Congelar fila de encabezado
    ws.freeze_panes = "A2"
 
 
def _agregar_hoja_resumen(wb, df: pd.DataFrame):
    """Agrega una hoja de resumen con estadísticas básicas del lote."""
    ws_res = wb.create_sheet("Resumen")
 
    font_titulo = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    fill_titulo = PatternFill("solid", fgColor="1F3864")
    font_label  = Font(name="Arial", bold=True, size=10)
    font_valor  = Font(name="Arial", size=10)
    alin_centro = Alignment(horizontal="center", vertical="center")
 
    # Título
    ws_res.merge_cells("A1:D1")
    celda_titulo = ws_res["A1"]
    celda_titulo.value = "RESUMEN DEL PROCESAMIENTO"
    celda_titulo.font = font_titulo
    celda_titulo.fill = fill_titulo
    celda_titulo.alignment = alin_centro
    ws_res.row_dimensions[1].height = 28
 
    datos_resumen = [
        ("Total de facturas procesadas", len(df)),
        ("Facturas con Nro. Factura", df["Nro. Factura"].notna().sum() if "Nro. Factura" in df.columns else "N/A"),
        ("Facturas con cliente identificado", df["Nombre Cliente"].notna().sum() if "Nombre Cliente" in df.columns else "N/A"),
        ("Facturas con total extraído", df["Total Facturación"].notna().sum() if "Total Facturación" in df.columns else "N/A"),
    ]
 
    if "Total Facturación" in df.columns:
        total_sum = pd.to_numeric(df["Total Facturación"], errors="coerce").sum()
        datos_resumen.append(("Suma total facturación", f"{total_sum:,.2f}"))
 
    for i, (label, valor) in enumerate(datos_resumen, start=3):
        ws_res.cell(row=i, column=1, value=label).font = font_label
        ws_res.cell(row=i, column=2, value=valor).font = font_valor
 
    ws_res.column_dimensions["A"].width = 38
    ws_res.column_dimensions["B"].width = 20
 
 
def _borde_fino() -> Border:
    lado = Side(style="thin", color=COLOR_BORDE)
    return Border(left=lado, right=lado, top=lado, bottom=lado)